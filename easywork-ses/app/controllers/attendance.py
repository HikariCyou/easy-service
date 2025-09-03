from calendar import monthrange
from datetime import date, datetime, timedelta, timezone
from io import BytesIO
from typing import List, Dict, Any, Optional
from decimal import Decimal
import pandas as pd
from starlette.responses import StreamingResponse
from tortoise.expressions import Q
from tortoise.transactions import in_transaction

from app.core.ctx import CTX_USER_ID
from app.models.attendance import DailyAttendance, MonthlyAttendance, WeeklyMood, MonthlyAttendanceLog
from app.models.contract import Contract
from app.models.personnel import Personnel
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl import load_workbook

from app.models.enums import AttendanceType, ApproveStatus, WeeklyMoodStatus
from app.schemas.attendance import (
    DailyAttendanceSchema,
    CreateDailyAttendanceSchema, 
    UpdateDailyAttendanceSchema,
    AttendanceStatsSchema,
)
from app.utils.common import to_hhmm


class AttendanceController:
    """勤怠管理コントローラー"""

    def __init__(self):
        pass

    async def create_daily_attendance(self, attendance_data: CreateDailyAttendanceSchema) -> DailyAttendance:
        """日次出勤記録を作成"""
        user_id = CTX_USER_ID.get()

        personnel = await Personnel.get_or_none(user_id=user_id)
        if not personnel:
            raise Exception("指定されたユーザーの要員情報が見つかりません")


        current_contract = await Contract.filter(
            personnel=personnel,
            contract_end_date__gte=attendance_data.work_date
        ).prefetch_related('case', 'case__client_company').first()

        # 重複チェック
        existing = await DailyAttendance.get_or_none(
            user_id=user_id,
            # contract_id=current_contract.id,
            work_date=attendance_data.work_date
        )
        if existing:
            raise Exception(f"指定日（{attendance_data.work_date}）の出勤記録は既に存在します")

        # 出勤記録作成
        attendance = await DailyAttendance.create(
            user_id=user_id or attendance_data.user_id,
            contract_id=  current_contract.id if current_contract else None,
            work_date=attendance_data.work_date,
            start_time=attendance_data.start_time if attendance_data.start_time else None,
            end_time=attendance_data.end_time if attendance_data.end_time else None,
            lunch_break_minutes= attendance_data.lunch_break_minutes,
            evening_break_minutes = attendance_data.evening_break_minutes,
            other_break_minutes = attendance_data.other_break_minutes,
            attendance_type=attendance_data.attendance_type,
            remark=attendance_data.remark
        )

        return attendance

    async def update_daily_attendance(self, attendance_id: int, update_data: UpdateDailyAttendanceSchema) -> DailyAttendance:
        """日次出勤記録を更新"""
        attendance = await DailyAttendance.get_or_none(id=attendance_id)
        if not attendance:
            raise Exception("出勤記録が見つかりません")

        # 承認済みの記録は更新不可
        if attendance.approved_status == ApproveStatus.APPROVED:
            raise Exception("承認済みの出勤記録は更新できません")

        # 更新
        update_dict = update_data.dict(exclude_unset=True, exclude_none=True)
        for field, value in update_dict.items():
            setattr(attendance, field, value)

        await attendance.save()
        return attendance

    async def get_daily_attendance_list(
        self, 
        page: int = 1, 
        page_size: int = 20,
        search_params: Dict = None
    ) -> Dict[str, Any]:
        """日次出勤記録一覧取得"""
        if search_params is None:
            search_params = {}

        query = Q()
        
        # 検索条件
        if search_params.get('user_id'):
            query &= Q(user_id=search_params['user_id'])
        if search_params.get('contract_id'):
            query &= Q(contract_id=search_params['contract_id'])
        if search_params.get('start_date'):
            query &= Q(work_date__gte=search_params['start_date'])
        if search_params.get('end_date'):
            query &= Q(work_date__lte=search_params['end_date'])
        if search_params.get('attendance_type'):
            query &= Q(attendance_type=search_params['attendance_type'])
        # 日次考勤不再有审批状态，审批在月次层面进行
        # if search_params.get('is_approved') is not None:
        #     if search_params['is_approved']:
        #         query &= Q(approved_status=ApproveStatus.APPROVED)
        #     else:
        #         query &= Q(approved_status__in=[ApproveStatus.PENDING, ApproveStatus.REJECT, ApproveStatus.WITHDRAWN])

        # 総数取得
        total = await DailyAttendance.filter(query).count()

        # データ取得
        attendances = await DailyAttendance.filter(query).prefetch_related(
            'contract', 'contract__personnel'
        ).order_by('-work_date', '-created_at').limit(page_size).offset((page - 1) * page_size).all()

        # 変換
        items = []
        for attendance in attendances:
            attendance_dict = await attendance.to_dict()
            # 日付関連情報を追加
            attendance_dict['weekday'] = attendance.weekday
            attendance_dict['weekday_name'] = attendance.weekday_name_jp
            attendance_dict['is_weekend'] = attendance.is_weekend
            attendance_dict['total_break_minutes'] = attendance.total_break_minutes
            # 日次考勤不再有审批状态
            # attendance_dict['is_approved'] = attendance.approved_status == ApproveStatus.APPROVED
            attendance_dict['is_approved'] = False  # 审批在月次层面处理
            
            if attendance.contract:
                attendance_dict['contract'] = await attendance.contract.to_dict()
            items.append(attendance_dict)

        return {
            "items": items,
            "total": total,
            "page": page,
            "page_size": page_size
        }

    async def get_daily_attendance(self, attendance_id: int) -> Optional[DailyAttendanceSchema]:
        """日次出勤記録詳細取得"""
        attendance = await DailyAttendance.get_or_none(id=attendance_id).prefetch_related(
            'contract', 'contract__personnel'
        )
        if not attendance:
            return None

        attendance_dict = await attendance.to_dict()
        if attendance.contract:
            attendance_dict['contract'] = await attendance.contract.to_dict()

        return attendance_dict

    async def approve_attendances(self, attendance_ids: List[int], approved_by: int) -> List[DailyAttendance]:
        """出勤記録を承認"""
        approved_attendances = []

        async with in_transaction():
            for attendance_id in attendance_ids:
                attendance = await DailyAttendance.get_or_none(id=attendance_id)
                if attendance and attendance.approved_status != ApproveStatus.APPROVED:
                    attendance.approved_status = ApproveStatus.APPROVED
                    attendance.approved_at = datetime.now(timezone.utc)
                    attendance.approved_by = approved_by
                    await attendance.save()
                    approved_attendances.append(attendance)

        return approved_attendances

    async def calculate_monthly_attendance(self, contract_id: int, year_month: str) -> MonthlyAttendance:
        """月次出勤集計を計算"""
        # 既存の月次記録チェック
        monthly_attendance = await MonthlyAttendance.get_or_none(
            contract_id=contract_id,
            year_month=year_month
        )

        if not monthly_attendance:
            # 契約情報取得
            contract = await Contract.get_or_none(id=contract_id)
            if not contract:
                raise Exception("契約が見つかりません")

            # 月次記録作成
            monthly_attendance = await MonthlyAttendance.create(
                user_id=contract.personnel.user_id if contract.personnel else None,
                contract_id=contract_id,
                year_month=year_month,
                total_working_hours=Decimal("0.0"),
                working_days=0,
                base_payment=Decimal("0"),
                total_payment=Decimal("0")
            )

        # 日次記録から集計計算
        await monthly_attendance.calculate_from_daily_records()
        return monthly_attendance

    async def get_monthly_attendance_list(
        self,
        page: int = 1,
        page_size: int = 20,
        search_params: Dict = None
    ) -> Dict[str, Any]:
        """月次出勤集計一覧取得"""
        if search_params is None:
            search_params = {}

        query = Q()

        # 検索条件
        if search_params.get('user_id'):
            query &= Q(user_id=search_params['user_id'])
        if search_params.get('contract_id'):
            query &= Q(contract_id=search_params['contract_id'])
        if search_params.get('year_month'):
            query &= Q(year_month=search_params['year_month'])
        if search_params.get('is_calculated') is not None:
            query &= Q(is_calculated=search_params['is_calculated'])
        if search_params.get('is_confirmed') is not None:
            query &= Q(is_confirmed=search_params['is_confirmed'])

        # 総数取得
        total = await MonthlyAttendance.filter(query).count()

        # データ取得
        monthly_attendances = await MonthlyAttendance.filter(query).prefetch_related(
            'contract', 'contract__personnel'
        ).order_by('-year_month').limit(page_size).offset((page - 1) * page_size).all()

        # 変換
        items = []
        for monthly in monthly_attendances:
            monthly_dict = await monthly.to_dict()
            if monthly.contract:
                monthly_dict['contract'] = await monthly.contract.to_dict()
            items.append(monthly_dict)

        return {
            "items": items,
            "total": total,
            "page": page,
            "page_size": page_size
        }

    async def confirm_monthly_attendance(self, monthly_id: int, confirmed_by: int) -> MonthlyAttendance:
        """月次出勤集計を確定"""
        monthly = await MonthlyAttendance.get_or_none(id=monthly_id)
        if not monthly:
            raise Exception("月次出勤記録が見つかりません")

        if not monthly.is_calculated:
            raise Exception("計算が完了していない月次記録は確定できません")

        monthly.is_confirmed = True
        monthly.confirmed_at = datetime.now(timezone.utc)
        monthly.confirmed_by = confirmed_by
        await monthly.save()

        return monthly

    async def get_attendance_calendar(self, contract_id: int, year_month: str) -> Dict[str, Any]:
        """出勤カレンダーデータ取得"""
        year, month = map(int, year_month.split('-'))
        
        # 月の日数を取得
        if month == 12:
            next_month_first_day = date(year + 1, 1, 1)
        else:
            next_month_first_day = date(year, month + 1, 1)
        
        last_day = (next_month_first_day - timedelta(days=1)).day

        # 契約情報取得
        contract = await Contract.get_or_none(id=contract_id).prefetch_related('personnel')
        if not contract:
            raise Exception("契約が見つかりません")

        # 該当月の出勤記録取得
        attendances = await DailyAttendance.filter(
            contract_id=contract_id,
            work_date__year=year,
            work_date__month=month
        ).all()

        # 日別データ作成
        attendance_dict = {att.work_date.day: att for att in attendances}
        calendar_data = []
        
        working_days_count = 0
        actual_working_days = 0
        total_hours = Decimal("0.0")

        for day in range(1, last_day + 1):
            current_date = date(year, month, day)
            attendance = attendance_dict.get(day)
            
            day_data = {
                "date": current_date,
                "day": day,
                "weekday": current_date.weekday(),  # 0=月曜日
                "is_weekend": current_date.weekday() >= 5,  # 土日判定
                "attendance": None
            }

            if attendance:
                day_data["attendance"] = await attendance.to_dict()
                actual_working_days += 1
                if attendance.actual_working_hours:
                    total_hours += attendance.actual_working_hours

            # 営業日かどうかの判定（土日以外）
            if not day_data["is_weekend"]:
                working_days_count += 1

            calendar_data.append(day_data)

        # 目標時間計算（契約の標準時間ベース）
        target_hours = contract.standard_working_hours or Decimal("160.0")

        return {
            "year_month": year_month,
            "contract_id": contract_id,
            "calendar_data": calendar_data,
            "working_days": working_days_count,
            "actual_working_days": actual_working_days,
            "remaining_days": working_days_count - actual_working_days,
            "total_hours": total_hours,
            "target_hours": target_hours
        }

    async def get_attendance_stats(self, search_params: Dict = None) -> AttendanceStatsSchema:
        """出勤統計取得"""
        if search_params is None:
            search_params = {}

        query = Q()
        
        # 統計期間の設定
        if search_params.get('start_date') and search_params.get('end_date'):
            query &= Q(work_date__gte=search_params['start_date'])
            query &= Q(work_date__lte=search_params['end_date'])
            period = f"{search_params['start_date']} - {search_params['end_date']}"
        else:
            # デフォルトは当月
            today = date.today()
            first_day = today.replace(day=1)
            query &= Q(work_date__gte=first_day, work_date__lte=today)
            period = today.strftime("%Y-%m")

        if search_params.get('user_id'):
            query &= Q(user_id=search_params['user_id'])
        if search_params.get('contract_id'):
            query &= Q(contract_id=search_params['contract_id'])

        # 基本統計
        all_records = await DailyAttendance.filter(query).all()
        total_records = len(all_records)
        approved_records = len([r for r in all_records if r.approved_status == ApproveStatus.APPROVED])
        pending_approval = total_records - approved_records

        # 出勤タイプ別統計
        type_stats = {}
        total_working_hours = Decimal("0.0")
        
        for record in all_records:
            att_type = record.attendance_type.value
            type_stats[att_type] = type_stats.get(att_type, 0) + 1
            if record.actual_working_hours:
                total_working_hours += record.actual_working_hours

        # 平均日次勤務時間
        working_records = [r for r in all_records if r.attendance_type == AttendanceType.NORMAL]
        average_daily_hours = None
        if working_records:
            working_hours_sum = sum(r.actual_working_hours or Decimal("0.0") for r in working_records)
            average_daily_hours = working_hours_sum / len(working_records)

        # 月次支払統計
        year_months = set()
        if search_params.get('start_date') and search_params.get('end_date'):
            start_date = search_params['start_date']
            end_date = search_params['end_date']
            current = start_date.replace(day=1)
            while current <= end_date:
                year_months.add(current.strftime("%Y-%m"))
                if current.month == 12:
                    current = current.replace(year=current.year + 1, month=1)
                else:
                    current = current.replace(month=current.month + 1)
        else:
            year_months.add(period)

        payment_query = Q(year_month__in=year_months)
        if search_params.get('user_id'):
            payment_query &= Q(user_id=search_params['user_id'])
        if search_params.get('contract_id'):
            payment_query &= Q(contract_id=search_params['contract_id'])

        monthly_records = await MonthlyAttendance.filter(payment_query).all()
        total_payments = sum(m.total_payment for m in monthly_records)
        confirmed_payments = sum(m.total_payment for m in monthly_records if m.is_confirmed)
        pending_payments = total_payments - confirmed_payments

        return {
            "period": period,
            "total_records": total_records,
            "approved_records": approved_records,
            "pending_approval": pending_approval,
            "normal_attendance": type_stats.get("NORMAL", 0),
            "paid_leave": type_stats.get("PAID_LEAVE", 0),
            "sick_leave": type_stats.get("SICK_LEAVE", 0),
            "absence": type_stats.get("ABSENCE", 0),
            "late_arrivals": type_stats.get("LATE", 0),
            "early_departures": type_stats.get("EARLY_LEAVE", 0),
            "total_working_hours": total_working_hours,
            "average_daily_hours": average_daily_hours,
            "total_payments": total_payments,
            "confirmed_payments": confirmed_payments,
            "pending_payments": pending_payments
        }

    async def bulk_create_attendances(self, attendances_data: List[CreateDailyAttendanceSchema]) -> List[DailyAttendance]:
        """一括出勤記録作成"""
        created_attendances = []

        async with in_transaction():
            for attendance_data in attendances_data:
                try:
                    attendance = await self.create_daily_attendance(attendance_data)
                    created_attendances.append(attendance)
                except Exception as e:
                    # 個別エラーは記録するが処理は継続
                    print(f"Failed to create attendance for {attendance_data.work_date}: {e}")
                    continue

        return created_attendances

    async def delete_daily_attendance(self, attendance_id: int) -> bool:
        """日次出勤記録削除"""
        attendance = await DailyAttendance.get_or_none(id=attendance_id)
        if not attendance:
            return False

        # 承認済みは削除不可
        if attendance.approved_status == ApproveStatus.APPROVED:
            raise Exception("承認済みの出勤記録は削除できません")

        await attendance.delete()
        return True

    async def get_user_attendance_data(
        self, 
        user_id: int, 
        period_type: str = "month",
        target_date: Optional[date] = None
    ) -> Dict[str, Any]:
        """
        ユーザーの勤怠情報を期間別で取得
        
        Args:
            user_id: システムユーザーID
            period_type: 期間タイプ（day/week/month）
            target_date: 基準日（未指定の場合は今日）
        """
        if not target_date:
            target_date = date.today()

        # Personnel情報を取得
        personnel = await Personnel.get_or_none(user_id=user_id)
        if not personnel:
            raise Exception("指定されたユーザーの要員情報が見つかりません")

        # 該当ユーザーの現在の契約を取得
        current_contracts = await Contract.filter(
            personnel=personnel,
            contract_end_date__gte=target_date
        ).prefetch_related('case', 'case__client_company').all()



        # 期間の計算
        if period_type == "day":
            start_date = target_date
            end_date = target_date
        elif period_type == "week":
            # 週の開始を月曜日とする
            days_since_monday = target_date.weekday()
            start_date = target_date - timedelta(days=days_since_monday)
            end_date = start_date + timedelta(days=6)
        elif period_type == "month":
            start_date = target_date.replace(day=1)
            if target_date.month == 12:
                next_month = target_date.replace(year=target_date.year + 1, month=1, day=1)
            else:
                next_month = target_date.replace(month=target_date.month + 1, day=1)
            end_date = next_month - timedelta(days=1)
        else:
            raise Exception("無効な期間タイプです。day/week/monthのいずれかを指定してください")

        # 勤怠データを取得
        attendance_records = await DailyAttendance.filter(
            user_id=user_id,
            work_date__gte=start_date,
            work_date__lte=end_date
        ).prefetch_related('contract', 'contract__case').order_by('work_date').all()

        # 勤怠データを整形
        attendance_data = []
        total_working_hours = 0.0
        approved_count = 0
        
        for record in attendance_records:
            record_dict = await record.to_dict()
            record_dict["actual_working_hours"] = record.actual_working_hours
            if record.contract:
                record_dict['contract'] = await record.contract.to_dict()
                if record.contract.case:
                    record_dict['contract']['case'] = await record.contract.case.to_dict()
            
            attendance_data.append(record_dict)
            
            if record.actual_working_hours:
                total_working_hours += float(record.actual_working_hours)

        # 契約情報を整形
        contracts_data = []
        for contract in current_contracts:
            contract_dict = await contract.to_dict()
            case = await contract.case
            if case:
                contract_dict['case'] = await case.to_dict()
                if contract.case.client_company:
                    contract_dict['case']['client_company'] = await contract.case.client_company.to_dict()
            contracts_data.append(contract_dict)

        # サマリー情報
        summary = {
            "total_working_hours": total_working_hours,
            "total_days": len(attendance_records),
        }

        return {
            "user_id": user_id,
            "personnel": await personnel.to_dict(),
            "contracts": contracts_data,
            "attendance_data": attendance_data,
            "period_info": {
                "period_type": period_type,
                "target_date": target_date.isoformat(),
                "start_date": start_date.isoformat(),
                "end_date": end_date.isoformat()
            },
            "summary": summary
        }

    async def get_staff_list_with_attendance(
        self,
        page: int = 1,
        page_size: int = 10,
        search_params: Dict = None
    ) -> Dict[str, Any]:
        """要員リスト取得（考勤データ汇总付き）"""
        if search_params is None:
            search_params = {}

        # 基础查询
        query = Q()
        
        # 搜索条件
        if search_params.get('keyword'):
            keyword = search_params['keyword']
            query &= (
                Q(name__icontains=keyword) | 
                Q(name_kana__icontains=keyword) | 
                Q(personnel_code__icontains=keyword)
            )
        
        if search_params.get('person_type'):
            query &= Q(person_type=search_params['person_type'])
        
        if search_params.get('employment_status'):
            query &= Q(employment_status=search_params['employment_status'])
        
        if search_params.get('is_active') is not None:
            query &= Q(is_active=search_params['is_active'])
        
        if search_params.get('nationality'):
            query &= Q(nationality__icontains=search_params['nationality'])

        # 总数统计
        total = await Personnel.filter(query).count()

        # 获取要员列表
        personnel_list = await Personnel.filter(query).prefetch_related(
            'contracts', 'contracts__case', 'contracts__case__client_company'
        ).order_by('name').limit(page_size).offset((page - 1) * page_size).all()

        # 处理数据
        items = []
        for personnel in personnel_list:
            personnel_dict = await personnel.to_dict()
            
            # 获取当前活跃的契约
            active_contracts = [c for c in personnel.contracts if c.contract_end_date >= date.today()]
            current_contract = active_contracts[0] if active_contracts else None
            
            # 会社信息
            company_info = None
            if personnel.person_type == 'bp_employee':
                # BP社员的会社信息
                company_info = {
                    "type": "bp_company",
                    "name": personnel.bp_company.company_name if hasattr(personnel, 'bp_company') and personnel.bp_company else "未設定"
                }
            elif personnel.person_type == 'employee':
                # 自社员工
                company_info = {
                    "type": "employee",
                    "name": "自社"
                }
            elif personnel.person_type == 'freelancer':
                # フリーランス
                company_info = {
                    "type": "freelancer", 
                    "name": "フリーランス"
                }

            # 案件信息
            case_info = None
            if current_contract and current_contract.case:
                case_info = {
                    "case_title": current_contract.case.title,
                    "client_company": current_contract.case.client_company.company_name if current_contract.case.client_company else "未設定",
                    "unit_price": float(current_contract.unit_price) if current_contract.unit_price else 0,
                    "contract_period": f"{current_contract.contract_start_date} 〜 {current_contract.contract_end_date}"
                }

            # 考勤数据汇总（指定月份或当月）
            target_year_month = search_params.get('year_month')
            attendance_summary = await self._calculate_personnel_attendance_summary(personnel.user_id, target_year_month)
            
            items.append({
                **personnel_dict,
                "company_info": company_info,
                "case_info": case_info,
                "attendance_summary": attendance_summary,
                "current_contract": await current_contract.to_dict() if current_contract else None
            })

        return {
            "items": items,
            "total": total,
            "page": page,
            "page_size": page_size
        }

    async def _calculate_personnel_attendance_summary(self, user_id: int, year_month: str = None) -> Dict[str, Any]:
        """计算要员的考勤数据汇总（指定月份或当月）"""
        if year_month:
            # 解析指定的年月
            year, month = map(int, year_month.split('-'))
            target_date = date(year, month, 1)
            # 获取该月的最后一天
            if month == 12:
                next_month = date(year + 1, 1, 1)
            else:
                next_month = date(year, month + 1, 1)
            last_day = next_month - timedelta(days=1)
            period_label = year_month
        else:
            # 使用当月
            today = date.today()
            target_date = today.replace(day=1)
            last_day = today
            period_label = today.strftime("%Y-%m")
        
        # 获取指定月份的考勤记录
        records = await DailyAttendance.filter(
            user_id=user_id,
            work_date__gte=target_date,
            work_date__lte=last_day
        ).all()
        
        total_days = len(records)
        total_hours = sum(record.actual_working_hours for record in records)
        paid_leave_days = len([r for r in records if r.attendance_type == AttendanceType.PAID_LEAVE])
        
        # 简单的残业计算（这里需要根据契约的标准工作时间来计算）
        # 假设每天标准8小时，月标准160小时
        standard_hours_per_day = 8.0
        expected_work_days = len([r for r in records if r.attendance_type == AttendanceType.NORMAL])
        expected_hours = expected_work_days * standard_hours_per_day
        overtime_hours = max(0.0, total_hours - expected_hours)
        
        return {
            "total_days": total_days,
            "total_hours": round(total_hours, 1),
            "overtime_hours": round(overtime_hours, 1),
            "paid_leave_days": paid_leave_days,
            "period": period_label
        }

    # === 心情管理相关方法 ===
    
    async def set_weekly_mood(self, user_id: int, mood_status: WeeklyMoodStatus, week_number:int = None,comment: str = None) -> WeeklyMood:
        """设置当前周的心情状态"""
        return await WeeklyMood.set_weekly_mood(user_id, mood_status, week_number, comment)
    
    async def get_current_week_mood(self, user_id: int) -> Optional[WeeklyMood]:
        """获取当前周的心情状态"""
        return await WeeklyMood.get_current_week_mood(user_id)
    
    async def get_mood_history(
        self, 
        user_id: int, 
        start_year: int = None, 
        start_week: int = None,
        limit: int = 12
    ) -> List[WeeklyMood]:
        """获取心情历史记录（最近几周）"""
        if start_year is None:
            from datetime import datetime
            today = datetime.now().date()
            start_year = today.year
            start_week = today.isocalendar()[1]
        
        moods = await WeeklyMood.filter(
            user_id=user_id,
            year__gte=start_year - 1,  # 跨年查询
            year__lte=start_year
        ).order_by('-year', '-week_number').limit(limit).all()
        
        return moods
    
    async def get_current_month_moods(self, user_id: int, year_month: str = None) -> Dict[str, Any]:
        """获取当前月份的4周心情数据"""
        from datetime import datetime, date, timedelta
        from calendar import monthrange
        
        if year_month:
            year, month = map(int, year_month.split('-'))
        else:
            today = date.today()
            year, month = today.year, today.month
            year_month = f"{year}-{month:02d}"
        
        # 获取该月第一天和最后一天
        first_day = date(year, month, 1)
        last_day = date(year, month, monthrange(year, month)[1])
        
        # 计算该月涉及的所有周（可能跨月）
        first_week = first_day.isocalendar()[1]
        last_week = last_day.isocalendar()[1]
        
        # 处理跨年情况
        if month == 1 and first_week > 50:  # 1月份可能包含去年的最后一周
            # 查询去年最后一周和今年的周
            moods_prev_year = await WeeklyMood.filter(
                user_id=user_id,
                year=year-1,
                week_number=first_week
            ).all()
            moods_current_year = await WeeklyMood.filter(
                user_id=user_id,
                year=year,
                week_number__lte=last_week
            ).all()
            all_moods = list(moods_prev_year) + list(moods_current_year)
        elif month == 12 and last_week < 10:  # 12月份可能包含明年第一周
            moods_current_year = await WeeklyMood.filter(
                user_id=user_id,
                year=year,
                week_number__gte=first_week
            ).all()
            moods_next_year = await WeeklyMood.filter(
                user_id=user_id,
                year=year+1,
                week_number=last_week
            ).all()
            all_moods = list(moods_current_year) + list(moods_next_year)
        else:
            # 正常情况，同一年内
            all_moods = await WeeklyMood.filter(
                user_id=user_id,
                year=year,
                week_number__gte=first_week,
                week_number__lte=last_week
            ).all()
        
        # 生成该月所有周的信息
        current_date = first_day
        month_weeks = []
        week_moods_map = {f"{mood.year}-W{mood.week_number:02d}": mood for mood in all_moods}
        
        while current_date <= last_day:
            week_year, week_number, weekday = current_date.isocalendar()
            week_key = f"{week_year}-W{week_number:02d}"
            
            # 计算这周在该月的日期范围
            week_start = current_date - timedelta(days=weekday-1)  # 周一
            week_end = week_start + timedelta(days=6)  # 周日
            
            # 计算该周在当前月的日期范围
            month_start_in_week = max(week_start, first_day)
            month_end_in_week = min(week_end, last_day)
            
            mood_record = week_moods_map.get(week_key)
            
            week_info = {
                "year": week_year,
                "week_number": week_number,
                "week_key": week_key,
                "week_start": week_start.isoformat(),
                "week_end": week_end.isoformat(),
                "month_days_in_week": f"{month_start_in_week.isoformat()} ~ {month_end_in_week.isoformat()}",
                "days_count_in_month": (month_end_in_week - month_start_in_week).days + 1,
                "mood_data": {
                    "status": mood_record.mood_status if mood_record else None,
                    "emoji": mood_record.mood_emoji if mood_record else "❓",
                    "score": mood_record.mood_score if mood_record else None,
                    "comment": mood_record.comment if mood_record else None,
                    "recorded_at": mood_record.created_at.isoformat() if mood_record else None,
                    "is_recorded": mood_record is not None
                } if mood_record else {
                    "status": None,
                    "emoji": "❓",
                    "score": None,
                    "comment": None,
                    "recorded_at": None,
                    "is_recorded": False
                }
            }
            
            month_weeks.append(week_info)
            
            # 移动到下一周的同一天（通常是周一）
            current_date += timedelta(days=7)
            if current_date.month != month and current_date.year == year:
                break
            elif current_date.year != year:
                break
        
        # 统计信息
        recorded_weeks = len([w for w in month_weeks if w["mood_data"]["is_recorded"]])
        total_score = sum([w["mood_data"]["score"] for w in month_weeks if w["mood_data"]["score"]])
        average_score = total_score / recorded_weeks if recorded_weeks > 0 else 0
        
        # 心情分布统计
        mood_distribution = {}
        for week in month_weeks:
            if week["mood_data"]["is_recorded"]:
                status = week["mood_data"]["status"]
                mood_distribution[status] = mood_distribution.get(status, 0) + 1
        
        return {
            "year_month": year_month,
            "total_weeks": len(month_weeks),
            "recorded_weeks": recorded_weeks,
            "unrecorded_weeks": len(month_weeks) - recorded_weeks,
            "average_score": round(average_score, 1) if average_score > 0 else None,
            "mood_distribution": mood_distribution,
            "weeks_data": month_weeks,
            "summary": {
                "completion_rate": f"{recorded_weeks}/{len(month_weeks)}",
                "most_common_mood": max(mood_distribution.items(), key=lambda x: x[1])[0] if mood_distribution else None
            }
        }
    
    async def get_team_mood_summary(self, year: int = None, week_number: int = None) -> Dict[str, Any]:
        """获取团队心情汇总（管理者用）"""
        if year is None or week_number is None:
            from datetime import datetime
            today = datetime.now().date()
            year = today.year
            week_number = today.isocalendar()[1]
        
        moods = await WeeklyMood.filter(year=year, week_number=week_number).all()
        
        # 统计各种心情状态的数量
        mood_counts = {}
        for mood in moods:
            status = mood.mood_status.value
            mood_counts[status] = mood_counts.get(status, 0) + 1
        
        total_count = len(moods)
        
        return {
            "year": year,
            "week_number": week_number,
            "total_responses": total_count,
            "mood_distribution": mood_counts,
            "mood_details": [await mood.to_dict() for mood in moods]
        }

    # === 月度考勤提交管理 ===
    
    async def get_or_create_monthly_attendance(self, user_id: int, year_month: str, contract_id: int = None) -> MonthlyAttendance:
        """获取或创建月度考勤记录"""
        # 确保year_month是字符串格式
        year_month_str = str(year_month) if year_month else ""
        
        # 获取用户的当前合同
        if not contract_id:
            personnel = await Personnel.get_or_none(user_id=user_id)
            if personnel:
                current_contracts = await Contract.filter(
                    personnel=personnel,
                    contract_end_date__gte=date.today()
                ).all()
                if current_contracts:
                    contract_id = current_contracts[0].id
        
        monthly, created = await MonthlyAttendance.get_or_create(
            user_id=user_id,
            year_month=year_month_str,
            defaults={
                "contract_id": contract_id,
                "status": ApproveStatus.DRAFT
            }
        )
        
        return monthly
    
    async def _log_monthly_operation(self, monthly_id: int, operation: str, operator_id: int, 
                                   from_status: str = None, to_status: str = None, remark: str = None):
        """记录月次考勤操作日志"""
        await MonthlyAttendanceLog.create(
            monthly_id=monthly_id,
            operation=operation,
            operator_id=operator_id,
            from_status=from_status,
            to_status=to_status,
            remark=remark
        )
    
    async def submit_monthly_attendance(self, user_id: int, year_month: str, remark: str = None) -> MonthlyAttendance:
        """智能提交月次考勤（根据状态处理）"""
        monthly, created = await MonthlyAttendance.get_or_create(
            user_id=user_id,
            year_month=str(year_month),
            defaults={"status": ApproveStatus.DRAFT}
        )
        
        original_status = monthly.status
        
        # 检查是否可以提交/重新提交
        if monthly.status == ApproveStatus.APPROVED:
            raise ValueError("已批准的考勤记录不能修改")
        
        # 根据当前状态处理
        if monthly.status in [ApproveStatus.DRAFT, ApproveStatus.REJECTED, ApproveStatus.WITHDRAWN]:
            # 可以提交或重新提交
            monthly.status = ApproveStatus.PENDING
            monthly.submitted_at = datetime.now()
            if remark:
                monthly.submit_remark = remark
            
            # 记录操作日志
            await self._log_monthly_operation(
                monthly_id=monthly.id,
                operation="submit" if original_status == ApproveStatus.DRAFT else "resubmit",
                operator_id=user_id,
                from_status=original_status,
                to_status=ApproveStatus.PENDING,
                remark=remark
            )
        elif monthly.status == ApproveStatus.PENDING:
            # 已在审批中，更新备注即可
            if remark:
                monthly.submit_remark = remark
                # 记录备注更新日志
                await self._log_monthly_operation(
                    monthly_id=monthly.id,
                    operation="update_remark",
                    operator_id=user_id,
                    from_status=original_status,
                    to_status=monthly.status,
                    remark=remark
                )
        
        await monthly.save()
        return monthly
    
    async def get_monthly_attendance_detail(self, monthly_id: int) -> Dict[str, Any]:
        """获取月次考勤详细数据（包含用户信息和日次记录）"""
        # 获取月度考勤记录
        monthly = await MonthlyAttendance.get_or_none(id=monthly_id)
        if not monthly:
            return None
        
        # 获取提交用户的要员信息
        personnel = await Personnel.get_or_none(user_id=monthly.user_id)
        user_info = None
        if personnel:
            user_info = await personnel.to_dict()
        
        # 获取该月的日次考勤记录
        year_month_str = str(monthly.year_month)
        if "-" not in year_month_str:
            raise ValueError(f"Invalid year_month format: {monthly.year_month}")
        year, month = map(int, year_month_str.split("-"))
        
        # 计算该月的开始和结束日期
        from calendar import monthrange
        start_date = date(year, month, 1)
        end_date = date(year, month, monthrange(year, month)[1])
        
        daily_records = await DailyAttendance.filter(
            user_id=monthly.user_id,
            work_date__gte=start_date,
            work_date__lte=end_date
        ).prefetch_related('contract').order_by('work_date').all()
        
        # 计算汇总统计
        total_working_hours = 0.0
        working_days = 0
        paid_leave_days = 0
        absence_days = 0
        
        daily_list = []
        for record in daily_records:
            # 缓存合同信息用于加班时间计算
            if hasattr(record, 'contract'):
                record._contract_cached = record.contract
                
            hours = record.actual_working_hours or 0.0
            total_working_hours += float(hours)
            
            if record.attendance_type == AttendanceType.NORMAL:
                working_days += 1
            elif record.attendance_type == AttendanceType.PAID_LEAVE:
                paid_leave_days += 1
            elif record.attendance_type == AttendanceType.ABSENCE:
                absence_days += 1
            
            # 直接使用to_dict，不要手动转换
            daily_list.append(await record.to_dict())
        
        # 组装返回数据
        result = {
            "monthly_info": await monthly.to_dict(),
            "user_info": user_info,
            "daily_records": daily_list,
            "summary": {
                "total_working_hours": round(total_working_hours, 2),
                "working_days": working_days,
                "paid_leave_days": paid_leave_days,
                "absence_days": absence_days,
                "total_days": len(daily_records)
            }
        }
        
        return result
    
    async def withdraw_monthly_attendance(self, monthly_id: int, user_id: int) -> MonthlyAttendance:
        """撤回月度考勤提交（简化参数）"""
        monthly = await MonthlyAttendance.get_or_none(id=monthly_id)
        
        if not monthly:
            raise Exception("月度考勤记录不存在")
        
        # 权限检查：只能撤回自己的记录
        if monthly.user_id != user_id:
            raise Exception("只能撤回自己的考勤记录")
        
        original_status = monthly.status
        await monthly.withdraw()
        
        # 记录操作日志
        await self._log_monthly_operation(
            monthly_id=monthly_id,
            operation="withdraw",
            operator_id=user_id,
            from_status=original_status,
            to_status=ApproveStatus.WITHDRAWN,
            remark="用户主动撤回"
        )
        
        return monthly
    
    async def approve_monthly_attendance(self, monthly_id: int, approved_by: int, remark: str = None) -> MonthlyAttendance:
        """批准月度考勤（管理者操作）"""
        monthly = await MonthlyAttendance.get_or_none(id=monthly_id)
        
        if not monthly:
            raise Exception("月度考勤记录不存在")
        
        original_status = monthly.status
        await monthly.approve(approved_by=approved_by, approve_remark=remark)
        
        # 记录操作日志
        await self._log_monthly_operation(
            monthly_id=monthly_id,
            operation="approve",
            operator_id=approved_by,
            from_status=original_status,
            to_status=ApproveStatus.APPROVED,
            remark=remark
        )
        
        return monthly
    
    async def reject_monthly_attendance(self, monthly_id: int, approved_by: int, remark: str = None) -> MonthlyAttendance:
        """拒绝月度考勤（管理者操作）"""
        monthly = await MonthlyAttendance.get_or_none(id=monthly_id)
        
        if not monthly:
            raise Exception("月度考勤记录不存在")
        
        original_status = monthly.status
        await monthly.reject(approved_by=approved_by, approve_remark=remark)
        
        # 记录操作日志
        await self._log_monthly_operation(
            monthly_id=monthly_id,
            operation="reject",
            operator_id=approved_by,
            from_status=original_status,
            to_status=ApproveStatus.REJECTED,
            remark=remark
        )
        
        return monthly
    
    async def get_monthly_attendance_status(self, user_id: int, year_month: str) -> Dict[str, Any]:
        """获取月度考勤状态和统计（自动创建记录）"""
        # 确保year_month是字符串格式
        year_month_str = str(year_month) if year_month else ""
        
        # 获取或创建月度记录
        monthly, created = await MonthlyAttendance.get_or_create(
            user_id=user_id,
            year_month=year_month_str,
            defaults={
                "status": ApproveStatus.DRAFT
            }
        )
        
        monthly_dict = await monthly.to_dict()
        
        # 如果有快照数据，使用快照；否则实时计算
        if monthly.snapshot_data:
            stats = monthly.snapshot_data
        else:
            daily_records = await monthly.get_daily_records()
            stats = {
                "total_working_hours": sum(record.actual_working_hours for record in daily_records),
                "working_days": len([r for r in daily_records if r.attendance_type == AttendanceType.NORMAL]),
                "paid_leave_days": len([r for r in daily_records if r.attendance_type == AttendanceType.PAID_LEAVE]),
                "absence_days": len([r for r in daily_records if r.attendance_type == AttendanceType.ABSENCE])
            }
        
        monthly_dict.update({
            "statistics": stats,
            "can_edit": monthly.status in [ApproveStatus.DRAFT, ApproveStatus.WITHDRAWN, ApproveStatus.REJECTED],
            "can_submit": monthly.status in [ApproveStatus.DRAFT, ApproveStatus.WITHDRAWN, ApproveStatus.REJECTED],
            "can_withdraw": monthly.status == ApproveStatus.PENDING
        })
        
        return monthly_dict

    async def get_attendance_data_for_export(self, user_id: int, year_month: str,
                                             include_summary: bool) -> StreamingResponse:
        year, month = map(int, year_month.split('-'))
        from calendar import monthrange
        from datetime import date
        from tortoise.expressions import Q

        start_date = date(year, month, 1)
        _, last_day = monthrange(year, month)
        end_date = date(year, month, last_day)

        q = Q()
        if user_id:
            q &= Q(user_id=user_id)
        q &= Q(work_date__gte=start_date) & Q(work_date__lte=end_date)

        attendances = await DailyAttendance.filter(q).order_by('work_date').prefetch_related('contract').all()

        data_rows = []
        for att in attendances:
            start_time = to_hhmm(att.start_time)
            end_time = to_hhmm(att.end_time)
            lunch_break = f"{att.lunch_break_minutes // 60:02d}:{att.lunch_break_minutes % 60:02d}" if att.lunch_break_minutes > 0 else ""
            evening_break = f"{att.evening_break_minutes // 60:02d}:{att.evening_break_minutes % 60:02d}" if att.evening_break_minutes > 0 else ""
            actual_hours = att.actual_working_hours
            working_hours = f"{int(actual_hours):02d}:{int((actual_hours % 1) * 60):02d}" if actual_hours > 0 else ""
            content = att.remark or ''
            # 添加出勤类型和日文星期
            attendance_type = att.attendance_type if hasattr(att, 'attendance_type') else AttendanceType.NORMAL
            
            # 日文星期映射
            weekday_jp = {
                'Monday': '月', 'Tuesday': '火', 'Wednesday': '水',
                'Thursday': '木', 'Friday': '金', 'Saturday': '土', 'Sunday': '日'
            }
            weekday = weekday_jp.get(att.work_date.strftime('%A'), '')
            
            data_rows.append([
                att.work_date.strftime('%m/%d'),  # 日本式日期格式
                weekday,                          # 日文星期
                start_time,
                end_time,
                lunch_break,
                evening_break,
                working_hours,
                attendance_type,  # 出勤区分
                content
            ])

        # サマリー計算（actual_working_hoursはpropertyなので手動計算）
        total_hours = sum(att.actual_working_hours for att in attendances)
        working_days = len([a for a in attendances if a.attendance_type == AttendanceType.NORMAL])
        paid_leave_days = len([a for a in attendances if a.attendance_type == AttendanceType.PAID_LEAVE])
        business_days = len(set(a.work_date for a in attendances))

        # 获取真实的用户信息
        from app.models.personnel import Personnel
        from app.models.case import Case
        
        name = "不明"
        department = "不明"
        project = ""
        
        if attendances:
            # 从第一个考勤记录获取用户ID
            first_attendance = attendances[0]
            personnel = await Personnel.get_or_none(user_id=first_attendance.user_id)
            if personnel:
                name = personnel.name
                # 从合同获取案件信息
                if hasattr(first_attendance, 'contract') and first_attendance.contract:
                    contract = first_attendance.contract
                    case = await Case.get_or_none(id=contract.case_id)
                    if case:
                        project = case.title or ""
                        # 可以从case获取部门信息，如果有的话
                        # department = case.department or "不明"
        title_month = year_month or start_date.strftime('%Y-%m')

        df = pd.DataFrame(data_rows,
                          columns=['日付', '曜日', '開始時刻', '退勤時刻', '昼休憩', '夜休憩', '作業時間', '出勤区分', '備考'])

        # 4. 写入 Excel带美化格式
        from openpyxl.styles import Border, Side, PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter
        
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='勤怠表', startrow=7)  # 数据从第8行开始
            ws = writer.sheets['勤怠表']

            # 设置表格样式
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # 标题样式
            header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True, size=11)
            
            # 日本式主标题（更简洁）
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)
            title_cell = ws.cell(row=1, column=1, value=f"{title_month}月 勤怠表")
            title_cell.font = Font(size=14, bold=True)
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[1].height = 30

            # 美化的基本信息区域
            info_font = Font(size=10)
            info_bold_font = Font(size=10, bold=True)
            
            # 第一行：氏名区域（合并单元格）
            ws.merge_cells('A3:B3')
            ws['A3'] = '氏名'
            ws['A3'].font = info_bold_font
            ws['A3'].alignment = Alignment(horizontal='center', vertical='center')
            ws['A3'].fill = PatternFill(start_color='E8E8E8', end_color='E8E8E8', fill_type='solid')
            
            ws.merge_cells('C3:F3')
            ws['C3'] = name
            ws['C3'].font = info_font
            ws['C3'].alignment = Alignment(horizontal='left', vertical='center')
            
            # 所属区域
            ws.merge_cells('G3:H3')
            ws['G3'] = '所属'
            ws['G3'].font = info_bold_font
            ws['G3'].alignment = Alignment(horizontal='center', vertical='center')
            ws['G3'].fill = PatternFill(start_color='E8E8E8', end_color='E8E8E8', fill_type='solid')
            
            ws['I3'] = department if department != '不明' else ''
            ws['I3'].font = info_font
            ws['I3'].alignment = Alignment(horizontal='left', vertical='center')
            
            # 第二行：案件名区域（合并单元格）
            ws.merge_cells('A4:B4')
            ws['A4'] = '案件名'
            ws['A4'].font = info_bold_font
            ws['A4'].alignment = Alignment(horizontal='center', vertical='center')
            ws['A4'].fill = PatternFill(start_color='E8E8E8', end_color='E8E8E8', fill_type='solid')
            
            ws.merge_cells('C4:I4')
            ws['C4'] = project
            ws['C4'].font = info_font
            ws['C4'].alignment = Alignment(horizontal='left', vertical='center')
            
            # 给信息区域加边框
            for row in [3, 4]:
                for col in range(1, 10):
                    ws.cell(row=row, column=col).border = thin_border
                    
            ws.row_dimensions[3].height = 22
            ws.row_dimensions[4].height = 22

            # 日本式表头（双线边框）
            thick_border = Border(
                left=Side(style='medium'),
                right=Side(style='medium'),
                top=Side(style='medium'),
                bottom=Side(style='medium')
            )
            
            for col in range(1, 10):  # 9列
                header_cell = ws.cell(row=8, column=col)
                header_cell.fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')  # 浅灰色
                header_cell.font = Font(bold=True, size=9)
                header_cell.alignment = Alignment(horizontal='center', vertical='center')
                header_cell.border = thick_border
                
            ws.row_dimensions[8].height = 22
            
            # 数据行（细边框，交替背景）
            light_fill = PatternFill(start_color='F8F8F8', end_color='F8F8F8', fill_type='solid')
            
            for row in range(9, 9 + len(data_rows)):
                ws.row_dimensions[row].height = 18
                # 偶数行添加浅色背景
                is_even_row = (row - 9) % 2 == 1
                
                for col in range(1, 10):
                    cell = ws.cell(row=row, column=col)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.font = Font(size=9)
                    
                    if is_even_row:
                        cell.fill = light_fill
                    
                    # 特殊列的对齐方式
                    if col == 9:  # 备注列左对齐
                        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    elif col == 8:  # 出勤区分列小字体
                        cell.font = Font(size=8)
                        
            # 日本式Excel列宽设置
            column_widths = [6, 4, 8, 8, 6, 6, 8, 10, 25]  # 更紧凑的列宽
            for i, width in enumerate(column_widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = width

            # 美化的汇总区域（左右分布式布局）
            summary_row = 9 + len(data_rows) + 2  # 空一行
            summary_fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
            summary_font = Font(bold=True, size=11)
            
            total_hours_formatted = f"{int(total_hours):02d}:{int((total_hours % 1) * 60):02d}"
            
            # 左侧：合计标题（合并3列）
            ws.merge_cells(start_row=summary_row, start_column=1, end_row=summary_row + 1, end_column=3)
            total_title = ws.cell(row=summary_row, column=1, value='合　計')
            total_title.font = Font(bold=True, size=14)
            total_title.alignment = Alignment(horizontal='center', vertical='center')
            total_title.fill = summary_fill
            
            # 为合计区域的每个单元格添加边框（使用细边框）
            for row in range(summary_row, summary_row + 2):
                for col in range(1, 4):
                    cell = ws.cell(row=row, column=col)
                    cell.border = thin_border
            
            # 右侧统计信息区域（6列，分两行显示）
            stats_titles = ['出勤日数', '有給取得', '総稼働時間']
            stats_values = [f'{working_days}日', f'{paid_leave_days}日', total_hours_formatted]
            
            # 上行：统计项目名称 - 先设置样式，再合并单元格
            for i, title in enumerate(stats_titles):
                col = 4 + i * 2
                # 先为要合并的单元格设置样式
                for c in range(col, col + 2):
                    title_cell = ws.cell(row=summary_row, column=c)
                    title_cell.font = Font(bold=True, size=9)
                    title_cell.alignment = Alignment(horizontal='center', vertical='center')
                    title_cell.fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
                    title_cell.border = thin_border
                
                # 然后设置值并合并
                ws.cell(row=summary_row, column=col, value=title)
                ws.merge_cells(start_row=summary_row, start_column=col, end_row=summary_row, end_column=col + 1)
                
            # 下行：统计数值 - 先设置样式，再合并单元格
            for i, value in enumerate(stats_values):
                col = 4 + i * 2
                # 先为要合并的单元格设置样式
                for c in range(col, col + 2):
                    value_cell = ws.cell(row=summary_row + 1, column=c)
                    value_cell.font = Font(bold=True, size=11)
                    value_cell.alignment = Alignment(horizontal='center', vertical='center')
                    value_cell.fill = summary_fill
                    value_cell.border = thin_border
                
                # 然后设置值并合并
                ws.cell(row=summary_row + 1, column=col, value=value)
                ws.merge_cells(start_row=summary_row + 1, start_column=col, end_row=summary_row + 1, end_column=col + 1)
                
            # 设置汇总区域行高
            ws.row_dimensions[summary_row].height = 25
            ws.row_dimensions[summary_row + 1].height = 25
            
            # 为整个报表添加蓝色外边框
            blue_border = Border(
                left=Side(style='medium', color='1f497d'),
                right=Side(style='medium', color='1f497d'),
                top=Side(style='medium', color='1f497d'),
                bottom=Side(style='medium', color='1f497d')
            )
            
            # 确定报表边界
            max_row = summary_row + 1
            max_col = 9
            
            # 添加外边框
            # 顶部边框
            for col in range(1, max_col + 1):
                ws.cell(row=1, column=col).border = Border(
                    top=Side(style='medium', color='1f497d'),
                    left=ws.cell(row=1, column=col).border.left,
                    right=ws.cell(row=1, column=col).border.right,
                    bottom=ws.cell(row=1, column=col).border.bottom
                )
            
            # 底部边框
            for col in range(1, max_col + 1):
                ws.cell(row=max_row, column=col).border = Border(
                    bottom=Side(style='medium', color='1f497d'),
                    left=ws.cell(row=max_row, column=col).border.left,
                    right=ws.cell(row=max_row, column=col).border.right,
                    top=ws.cell(row=max_row, column=col).border.top
                )
            
            # 左侧边框
            for row in range(1, max_row + 1):
                ws.cell(row=row, column=1).border = Border(
                    left=Side(style='medium', color='1f497d'),
                    top=ws.cell(row=row, column=1).border.top,
                    right=ws.cell(row=row, column=1).border.right,
                    bottom=ws.cell(row=row, column=1).border.bottom
                )
            
            # 右侧边框
            for row in range(1, max_row + 1):
                ws.cell(row=row, column=max_col).border = Border(
                    right=Side(style='medium', color='1f497d'),
                    left=ws.cell(row=row, column=max_col).border.left,
                    top=ws.cell(row=row, column=max_col).border.top,
                    bottom=ws.cell(row=row, column=max_col).border.bottom
                )

        output.seek(0)
        headers = {
            'Content-Disposition': 'attachment; filename="attendance_export.xlsx"'
        }
        return StreamingResponse(output,
                                 media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                                 headers=headers)


# グローバルインスタンス
attendance_controller = AttendanceController()